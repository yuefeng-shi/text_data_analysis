import re
import argparse
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from sklearn.metrics import accuracy_score, f1_score
from openpyxl.styles import Alignment

warnings.filterwarnings("ignore")

# =============================================================================
# CONSTANTS
# =============================================================================
SIZE_LIMIT_MULTIPLIER     = 1.5
GOLD_TOLERANCE_MULTIPLIER = 5.0

# Fallback range for RNRMSE_by_Range when gold range within a data_size is 0.
PCT_GOLD_RANGE_FALLBACK = 0.6  # = 0.8 - 0.2


# =============================================================================
# MODEL REGISTRY
# =============================================================================
MODEL_REGISTRY = {
    # ── OpenAI ──────────────────────────────────────────────────────────────
    "gpt-4.1":                       "openai/gpt-4.1",
    "gpt-4.1-mini":                  "openai/gpt-4.1-mini",
    "gpt-5":                         "openai/gpt-5",
    "gpt-5-mini":                    "openai/gpt-5-mini",
    "gpt-5.1":                       "openai/gpt-5.1",
    "gpt-5.4":                       "openai/gpt-5.4",
    "gpt-5.4-mini":                  "openai/gpt-5.4-mini",
    # ── Anthropic ───────────────────────────────────────────────────────────
    "claude-4":                      "anthropic/claude-sonnet-4",
    # ── Google Gemini ────────────────────────────────────────────────────────
    "gemini-2.5-flash":              "google/gemini-2.5-flash",
    "gemini-2.5-pro":                "google/gemini-2.5-pro",
    "gemini-3-flash":                "google/gemini-3-flash-preview",
    "gemini-3-pro":                  "google/gemini-3-pro-preview",
    "gemini-3.1-flash-lite":         "google/gemini-3.1-flash-lite-preview",
    "gemini-3.1-pro":                "google/gemini-3.1-pro-preview",
    # ── xAI Grok ────────────────────────────────────────────────────────────
    "grok-4":                        "x-ai/grok-4",
    "grok-4-fast":                   "x-ai/grok-4-fast",
    # ── DeepSeek ─────────────────────────────────────────────────────────────
    "deepseek-v3-0324":              "deepseek/deepseek-chat-v3-0324",
    "deepseek-v3.1":                 "deepseek/deepseek-chat-v3.1",
    "deepseek-v3.1-terminus":        "deepseek/deepseek-v3.1-terminus",
    # ── Meta LLaMA ───────────────────────────────────────────────────────────
    "llama-3.1-8b-instruct":         "meta-llama/llama-3.1-8b-instruct",
    "llama-3.2-3b-instruct":         "meta-llama/llama-3.2-3b-instruct",
    "llama-3.3-70b-instruct":        "meta-llama/llama-3.3-70b-instruct",
    "llama-4-scout":                 "meta-llama/llama-4-scout",
    # ── Qwen 2.5 ─────────────────────────────────────────────────────────────
    "qwen-2.5-72b-instruct":         "qwen/qwen-2.5-72b-instruct",
    "qwen-2.5-7b-instruct":          "qwen/qwen-2.5-7b-instruct",
    "qwen-2.5-3b-instruct":          "qwen/qwen-2.5-3b-instruct",
    "qwen2.5-coder-32b-instruct":    "qwen/qwen-2.5-coder-32b-instruct",
    "qwen2.5-coder-7b-instruct":     "qwen/qwen-2.5-coder-7b-instruct",
    # ── Qwen 3 ───────────────────────────────────────────────────────────────
    "qwen3-coder-480b-a35b":         "qwen/qwen3-coder-480b-a35b-instruct",
    "qwen3-235b-a22b":               "qwen/qwen3-235b-a22b",
    "qwen3-235b-a22b-2507-instruct": "qwen/qwen3-235b-a22b-2507",
    "qwen3-30b-a3b-instruct-2507":   "qwen/qwen3-30b-a3b-instruct-2507",
    "qwen3-next-80b-instruct":       "qwen/qwen3-next-80b-a3b-instruct",
    # ── Qwen 3.5 ─────────────────────────────────────────────────────────────
    "qwen3.5-9b":                    "qwen/qwen3.5-9b",
    "qwen3.5-27b":                   "qwen/qwen3.5-27b",
    "qwen3.5-122b-a10b":             "qwen/qwen3.5-122b-a10b",
    "qwen3.5-397b-a17b":             "qwen/qwen3.5-397b-a17b",
    "qwen3.5-flash":                 "qwen/qwen3.5-flash",
}

# =============================================================================
# MODEL SORT ORDER
# =============================================================================
MANUAL_MODEL_RANKS = {
    # OpenAI  (10-19)
    "gpt-5.4-mini":                  10,
    "gpt-5.4":                       11,
    "gpt-5.1":                       12,
    "gpt-5-mini":                    13,
    "gpt-5":                         14,
    "gpt-4.1-mini":                  15,
    "gpt-4.1":                       16,
    # Anthropic  (20-29)
    "claude-4":                      20,
    "claude-sonnet-4":               21,
    # Gemini  (30-39)
    "gemini-3.1-pro":                30,
    "gemini-3.1-flash-lite":         31,
    "gemini-3-pro":                  32,
    "gemini-3-flash":                33,
    "gemini-2.5-pro":                34,
    "gemini-2.5-flash":              35,
    # Grok  (40-49)
    "grok-4-fast":                   40,
    "grok-4":                        41,
    # LLaMA  (100-109)
    "llama-4-scout":                 100,
    "llama-3.3-70b-instruct":        101,
    "llama-3.1-8b-instruct":         102,
    "llama-3.2-3b-instruct":         103,
    # Qwen 3.5  (110-119)
    "qwen3.5-397b-a17b":             110,
    "qwen3.5-122b-a10b":             111,
    "qwen3.5-27b":                   112,
    "qwen3.5-9b":                    113,
    "qwen3.5-flash":                 114,
    # Qwen 3  (120-129)
    "qwen3-coder-480b-a35b":         120,
    "qwen3-235b-a22b-2507-instruct": 121,
    "qwen3-235b-a22b":               122,
    "qwen3-next-80b-instruct":       123,
    "qwen3-30b-a3b-instruct-2507":   124,
    # Qwen 2.5  (130-139)
    "qwen2.5-coder-32b-instruct":    130,
    "qwen2.5-coder-7b-instruct":     131,
    "qwen-2.5-72b-instruct":         132,
    "qwen-2.5-7b-instruct":          133,
    "qwen-2.5-3b-instruct":          134,
    # DeepSeek  (200-209)
    "deepseek-v3.1-terminus":        200,
    "deepseek-v3.1":                 201,
    "deepseek-v3-0324":              202,
}

CUSTOM_MODEL_ORDER = list(MANUAL_MODEL_RANKS.keys())

# =============================================================================
# MODEL TYPE CLASSIFICATION
# =============================================================================
MODEL_TYPES = {
    "Closed-source Models": ["gpt", "gemini", "grok", "claude"],
    "Open-weight Large":    ["llama-3.3-70b", "qwen3.5-397b", "qwen3.5-122b"],
    "Open-weight Models":   ["llama", "qwen", "deepseek", "mixtral", "gemma"],
}

TYPE_ORDER_MAP = {
    "Closed-source Models": 0,
    "Open-weight Large":    1,
    "Open-weight Models":   2,
    "Other Models":         3,
}

# =============================================================================
# DATASET CATEGORIES  (te_ / ste_ prefixes)
# =============================================================================
DATASET_CATEGORIES = {
    "Single-Label Classification": [
        "te-sentiment_analysis",
        "te-hate_speech_detection",
        "te-offensive_language_detection",
        "ste-tweethate",
    ],
    "Multi-Label Classification": [
        "ste-emotion",
        "ste-tweettopic",
    ],
    "Targeted Classification": [
        "ste-sentiment_analysis",
        "te-stance_detection",
    ],
}

CAT_ORDER = ["Single-Label Classification", "Multi-Label Classification", "Targeted Classification"]

# =============================================================================
# QUESTION TYPE REGISTRY
# =============================================================================
QUESTION_TYPES = {
    "existence":   ["yes/no"],
    "count":       ["number"],
    "calculation": ["number", "percentage"],
    "comparison":  ["yes/no"],
    "statistics":  ["yes/no", "number", "percentage"],
}

# =============================================================================
# TASK-TYPE BREAKDOWN TABLE
# =============================================================================
_TASK_ROW_SPECS = [
    ("Existence (F1)",        "existence",  "yes/no",     "F1-Macro"),
    ("Comparative (F1)",      "compar",     "yes/no",     "F1-Macro"),
    ("Calculation (RNRMSE)",  "calculat",   "percentage", "RNRMSE_by_Range"),
    ("Calculation (NRMSE)",   "calculat",   "number",     "NRMSE_by_Size"),
    ("Count (NRMSE)",         "count",      "number",     "NRMSE_by_Size"),
]

_TASK_MODEL_GROUPS = {
    "Closed-source": ["gpt", "gemini", "grok", "claude"],
    "Open-weight":   ["llama", "qwen", "deepseek", "mixtral"],
}

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def to_numeric(value) -> float:
    """Convert a string / number to float; return nan on failure."""
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return np.nan
    v = value.strip()
    if not v:
        return np.nan
    if "%" in v:
        try:
            return float(v.replace("%", "")) / 100.0
        except ValueError:
            return np.nan
    try:
        return float(v)
    except ValueError:
        return np.nan


def parse_output(output: str, answer_type: str):
    """
    Parse a raw model output string into a typed answer value (optimized mode).
    Returns None when the output is a refusal, error, or unparseable.
    """
    if not isinstance(output, str) or not output.strip():
        return None
    lower = output.lower()

    _BLOCK = [
        "no_answer", "no answer", "i can't", "i cannot", "i am unable", "unable to",
        "cannot provide", "cannot fulfill", "cannot answer", "as an ai", "sorry",
        "apologize", "context too long", "context length", "response that",
        "to determine", "to calculate", "calculating", "steps to", "the calculation",
        "neutral%", "determine the total", "based on the",
        "error code", "status code", "error message", "traceback", "exception",
        "syntax error", "http error", "json decode error",
    ]
    if any(k in lower for k in _BLOCK):
        return None

    if answer_type == "yes/no":
        if "negative" in lower or re.search(r"\bno\b", lower):
            return "no"
        if "positive" in lower or re.search(r"\byes\b", lower):
            return "yes"
        return None

    if answer_type in ("number", "percentage"):
        pct = re.search(r"(-?\d+\.?\d*)\s*%", output)
        if answer_type == "percentage" and pct:
            try:
                return float(pct.group(1)) / 100.0
            except ValueError:
                return None
        sci = re.search(r"-?\d+\.?\d*[eE][+-]?\d+", output)
        if sci:
            try:
                return float(sci.group(0))
            except ValueError:
                pass
        num = re.search(r"-?\d+\.?\d*", output)
        if num:
            try:
                return float(num.group(0))
            except ValueError:
                return None
        return None

    return output


def get_model_sort_key(model_name: str) -> float:
    """Return a numeric sort rank for a model name (lower = shown first)."""
    m = str(model_name).lower()
    for key, rank in MANUAL_MODEL_RANKS.items():
        if key in m:
            return rank
    size_match = re.search(r"(\d+(\.\d+)?)b", m)
    size_val = float(size_match.group(1)) if size_match else 0
    return 1000 + max(0.0, 1000.0 - size_val)


def get_model_type(model_name: str) -> str:
    """Classify a model into Closed-source / Open-weight Large / Open-weight / Other."""
    m = str(model_name).lower()
    for kw in MODEL_TYPES["Closed-source Models"]:
        if kw in m:
            return "Closed-source Models"
    for kw in MODEL_TYPES["Open-weight Large"]:
        if kw in m:
            return "Open-weight Large"
    for kw in MODEL_TYPES["Open-weight Models"]:
        if kw in m:
            return "Open-weight Models"
    return "Other Models"


def get_dataset_category(dataset_name: str):
    """Map a dataset name to its evaluation category string (or None)."""
    d = str(dataset_name).lower()
    for cat, keywords in DATASET_CATEGORIES.items():
        if any(d.startswith(kw) for kw in keywords):
            return cat
    return None


# =============================================================================
# DATA TRANSFORMATION
# =============================================================================

def transform_wide_to_long(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert wide CSV format into long format.

    Required columns
    ----------------
    Dataset, Question-Type, Answer-type, Question

    Column patterns recognised
    --------------------------
    gold_{size}_{sample}        : gold labels; size and sample are integers.
                                  Any subset of sizes / sample numbers is valid.
    res_{size}_{sample}_{model} : model predictions; paired with matching gold column.

    Note: base_ columns in the input CSV are intentionally ignored.
          Baselines are always injected synthetically by add_baselines().
    """
    required_cols = ["Dataset", "Question-Type", "Answer-type", "Question"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns in CSV: {missing}")

    id_vars = [c for c in required_cols if c in df.columns]
    if "Question-Type-II" in df.columns:
        id_vars.append("Question-Type-II")
    if "Question Category" in df.columns:
        id_vars.append("Question Category")
    if "target" in df.columns:
        id_vars.append("target")

    # ── Discover gold columns: gold_{size}_{sample} ───────────────────────────
    # size and sample must both be integers; any subset of sizes/samples is valid.
    gold_map: dict[str, str] = {}   # key: "{size}_{sample}" → column name
    for col in df.columns:
        m = re.match(r"^gold_(\d+)_(\d+)$", col)
        if m:
            gold_map[f"{m.group(1)}_{m.group(2)}"] = col

    if not gold_map:
        print("  [WARN] No gold columns found. Expected pattern: gold_{size}_{sample}")
        return pd.DataFrame()

    melted: list[pd.DataFrame] = []

    # ── Model prediction columns: res_{size}_{sample}_{model_name} ───────────
    for col in df.columns:
        m = re.match(r"^res_(\d+)_(\d+)_(.*)", col)
        if not m:
            continue
        size, sample, model_name = m.group(1), m.group(2), m.group(3)
        key = f"{size}_{sample}"
        if key not in gold_map:
            continue
        sub = df[id_vars + [gold_map[key], col]].copy()
        sub = sub.rename(columns={gold_map[key]: "Gold_Label", col: "Model_Output"})
        sub["Model"]     = model_name
        sub["Data_Size"] = int(size)
        sub["Data_ID"]   = int(sample)
        melted.append(sub)


    if not melted:
        return pd.DataFrame()

    long_df = pd.concat(melted, ignore_index=True)
    long_df.dropna(subset=["Model_Output"], inplace=True)
    return long_df



def add_baselines(models_long_df: pd.DataFrame) -> pd.DataFrame:
    """
    Append synthetic constant-prediction baselines to the long DataFrame.
    Any base_ columns present in the original CSV are ignored — baselines
    are always generated fresh here so results are consistent regardless
    of input file contents.

    yes/no baselines    : base_yes, base_no
    percentage baselines: base_0per, base_50per, base_100per
    number baselines    : DYNAMIC — scaled to Data_Size per row
                          base_num_0    → always predict 0
                          base_num_half → always predict Data_Size // 2
                          base_num_full → always predict Data_Size
    """
    if models_long_df.empty:
        return models_long_df
    print("  Adding baseline models ...")

    _STATIC_BASELINES = [
        ("base_yes",    "yes/no",     "yes"),
        ("base_no",     "yes/no",     "no"),
        ("base_0per",   "percentage", "0%"),
        ("base_50per",  "percentage", "50%"),
        ("base_100per", "percentage", "100%"),
    ]

    _DYNAMIC_NUM_BASELINES = [
        ("base_num_0",    lambda ds: 0),
        ("base_num_half", lambda ds: ds // 2),
        ("base_num_full", lambda ds: ds),
    ]

    template_cols = [
        "Dataset", "Question Category", "Question-Type",
        "Answer-type", "Question", "Gold_Label", "Data_Size", "Data_ID",
    ]
    template_cols = [c for c in template_cols if c in models_long_df.columns]
    if "target" in models_long_df.columns:
        template_cols.append("target")

    template = models_long_df[template_cols].drop_duplicates().reset_index(drop=True)
    parts    = [models_long_df]

    # Static baselines
    for name, atype, value in _STATIC_BASELINES:
        rows = template[template["Answer-type"] == atype].copy()
        if not rows.empty:
            rows["Model"]        = name
            rows["Model_Output"] = value
            parts.append(rows)

    # Dynamic number baselines
    num_template = template[template["Answer-type"] == "number"].copy()
    if not num_template.empty:
        for name, size_fn in _DYNAMIC_NUM_BASELINES:
            rows = num_template.copy()
            rows["Model"]        = name
            rows["Model_Output"] = rows["Data_Size"].apply(size_fn)
            parts.append(rows)

    return pd.concat(parts, ignore_index=True)

# =============================================================================
# CORE METRIC CALCULATION
# Parsing mode : optimized  (hardcoded)
# Error mode   : strict     (hardcoded — invalid outputs receive worst-case penalty)
# =============================================================================

def _validate_parsed(parsed, answer_type: str, data_size: int, gold) -> bool:
    """Return True when the parsed value passes range / sanity checks."""
    if parsed is None:
        return False
    if answer_type == "yes/no":
        return parsed in ("yes", "no")
    if answer_type == "percentage":
        return isinstance(parsed, (int, float)) and 0 <= parsed <= 1
    if answer_type == "number":
        if not (isinstance(parsed, (int, float)) and parsed >= 0):
            return False
        gold_num = to_numeric(gold)
        if parsed <= data_size * SIZE_LIMIT_MULTIPLIER:
            return True
        if pd.notna(gold_num) and gold_num > 0 and parsed <= gold_num * GOLD_TOLERANCE_MULTIPLIER:
            return True
        return False
    return False


def calculate_metrics(
    group: pd.DataFrame,
    answer_type: str,
) -> pd.DataFrame:
    """
    Compute evaluation metrics for one (dataset, model, answer_type) group.
    Always uses optimized parsing and strict error mode.
    Returns a small DataFrame with columns [Metric, Value, Count, Total, Valid_Count].
    """
    total       = len(group)
    valid_count = 0
    y_true_cat, y_pred_cat = [], []
    y_true_num, y_pred_num, num_sizes, pct_ranges = [], [], [], []

    # Precompute gold range per data_size for RNRMSE_by_Range
    if answer_type == "percentage" and "Data_Size" in group.columns:
        _gold_by_size: dict = {}
        for _, _r in group.iterrows():
            _ds = int(_r.get("Data_Size", 0))
            _gv = to_numeric(_r["Gold_Label"])
            if pd.notna(_gv):
                _gold_by_size.setdefault(_ds, []).append(_gv)
        _size_ranges = {
            ds: (max(vs) - min(vs)) if len(vs) > 1 else PCT_GOLD_RANGE_FALLBACK
            for ds, vs in _gold_by_size.items()
        }
    else:
        _size_ranges = {}

    for _, row in group.iterrows():
        gold, output = row["Gold_Label"], row["Model_Output"]
        data_size    = int(row.get("Data_Size", 0))

        parsed   = parse_output(str(output), answer_type)
        is_valid = _validate_parsed(parsed, answer_type, data_size, gold)

        if is_valid:
            valid_count += 1
            if answer_type == "yes/no":
                y_true_cat.append(str(gold).lower().strip())
                y_pred_cat.append(parsed)
            else:
                y_true_num.append(to_numeric(gold))
                y_pred_num.append(parsed)
                if answer_type == "number":
                    num_sizes.append(data_size)
                elif answer_type == "percentage":
                    pct_ranges.append(_size_ranges.get(data_size, PCT_GOLD_RANGE_FALLBACK))

        else:
            # Strict penalty: assign worst-case opposite prediction
            if answer_type == "yes/no":
                gl = str(gold).lower().strip()
                y_true_cat.append(gl)
                y_pred_cat.append("no" if gl == "yes" else "yes")
            else:
                gn = to_numeric(gold)
                if pd.notna(gn):
                    if answer_type == "percentage":
                        y_true_num.append(gn)
                        y_pred_num.append(1.0 if gn <= 0.5 else 0.0)
                        pct_ranges.append(_size_ranges.get(data_size, PCT_GOLD_RANGE_FALLBACK))
                    elif answer_type == "number":
                        y_true_num.append(gn)
                        y_pred_num.append(float(data_size) if gn <= data_size / 2.0 else 0.0)
                        num_sizes.append(data_size)

    n_participation = len(y_true_cat) + len(y_true_num)
    if not n_participation:
        return pd.DataFrame()

    valid_rate = valid_count / total if total else 0
    results    = []

    fv_map = {
        "yes/no":     "Format_Validity_F1",
        "percentage": "Format_Validity_MSE",
        "number":     "Format_Validity_NRMSE",
    }
    if answer_type in fv_map:
        results.append({"Metric": fv_map[answer_type], "Value": valid_rate})

    if y_true_cat:
        results.append({
            "Metric": "F1-Macro",
            "Value":  f1_score(y_true_cat, y_pred_cat, average="macro",
                               labels=["yes", "no"], zero_division=0),
        })
        results.append({
            "Metric": "Accuracy",
            "Value":  accuracy_score(y_true_cat, y_pred_cat),
        })

    if y_true_num:
        if answer_type == "percentage":
            _ranges   = pct_ranges if pct_ranges else [PCT_GOLD_RANGE_FALLBACK] * len(y_true_num)
            rnrmse_sq = [((t - p) / r) ** 2 if r > 0 else 0.0
                         for t, p, r in zip(y_true_num, y_pred_num, _ranges)]
            results.append({
                "Metric": "RNRMSE_by_Range",
                "Value":  float(np.sqrt(np.mean(rnrmse_sq))),
            })
        elif answer_type == "number" and num_sizes:
            sq = [((t - p) / s) ** 2 if s > 0 else (0.0 if t == p else np.nan)
                  for t, p, s in zip(y_true_num, y_pred_num, num_sizes)]
            sq = [v for v in sq if not np.isnan(v)]
            results.append({
                "Metric": "NRMSE_by_Size",
                "Value":  float(np.sqrt(np.mean(sq))) if sq else np.nan,
            })

    df_res = pd.DataFrame(results)
    if not df_res.empty:
        df_res["Count"]       = n_participation
        df_res["Total"]       = total
        df_res["Valid_Count"] = valid_count
    return df_res


# =============================================================================
# PART 1 – HIERARCHICAL EVALUATION
# =============================================================================

def perform_hierarchical_evaluation(long_df: pd.DataFrame) -> pd.DataFrame:
    """
    Evaluate at four hierarchy levels.
    Parsing mode (optimized) and error mode (strict) are hardcoded.

    Levels
    ------
    Overall         – one row per dataset (sizes pooled)
    Overall(NoSize) – per (dataset, question-type); used for summary sheet only
    By Size         – per (dataset, size)
    By Instance     – per (dataset, size, instance-id)
    """
    if long_df.empty:
        return pd.DataFrame()

    levels = [
        {
            "name": "Overall",
            "cols": ["Dataset"],
            "fmt":  lambda r: r["Dataset"],
        },
        {
            "name": "Overall (No Size)",
            "cols": ["Dataset", "Question-Type"],
            "fmt":  lambda r: f"{r['Dataset']} - {r['Question-Type']}",
        },
        {
            "name": "By Size",
            "cols": ["Dataset", "Data_Size"],
            "fmt":  lambda r: f"{r['Dataset']} | size_{r['Data_Size']}",
        },
        {
            "name": "By Instance",
            "cols": ["Dataset", "Data_Size", "Data_ID"],
            "fmt":  lambda r: f"{r['Dataset']} | size_{r['Data_Size']} | instance_{r['Data_ID']}",
        },
    ]

    all_parts = []
    for level in levels:
        group_cols = level["cols"] + ["Model", "Answer-type"]
        parts      = []

        for name, grp in long_df.groupby(group_cols):
            if not isinstance(name, tuple):
                name = (name,)
            keys        = dict(zip(group_cols, name))
            answer_type = keys["Answer-type"]

            res = calculate_metrics(grp, answer_type)
            if not res.empty:
                for k, v in keys.items():
                    res[k] = v
                res["Hierarchy"]       = level["fmt"](pd.Series(keys))
                res["Hierarchy_Level"] = level["name"]
                parts.append(res)

        if parts:
            all_parts.append(pd.concat(parts, ignore_index=True))

    return pd.concat(all_parts, ignore_index=True) if all_parts else pd.DataFrame()


def save_error_report(long_df: pd.DataFrame, out_dir: Path) -> None:
    """Save error_summary.csv and error_detail_log.csv to out_dir."""
    print("  Generating error report ...")
    error_log  = []
    base_fields = ["Model", "Answer-type", "Question", "Gold_Label", "Model_Output", "Data_Size"]
    if "target" in long_df.columns:
        base_fields.append("target")

    for _, row in long_df.iterrows():
        output, atype = row["Model_Output"], row["Answer-type"]
        parsed = parse_output(str(output), atype)

        error_type = None
        if parsed is None:
            error_type = "Parsing Failure (Refusal / CoT / Blocklist)"
        elif atype == "percentage" and not (0 <= parsed <= 1):
            error_type = "Value Error (Percentage out of 0-1 range)"
        elif atype == "number":
            if parsed < 0:
                error_type = "Value Error (Negative number)"
            else:
                ds = int(row.get("Data_Size", 0))
                gn = to_numeric(row["Gold_Label"])
                ok = parsed <= ds * SIZE_LIMIT_MULTIPLIER or (
                     pd.notna(gn) and gn > 0 and parsed <= gn * GOLD_TOLERANCE_MULTIPLIER)
                if not ok:
                    error_type = (
                        f"Value Error (Hallucination: "
                        f">{SIZE_LIMIT_MULTIPLIER}x Size & "
                        f">{GOLD_TOLERANCE_MULTIPLIER}x Gold)"
                    )

        if error_type:
            entry = {c: row.get(c) for c in base_fields}
            entry["Error_Type"]   = error_type
            entry["Parsed_Value"] = parsed
            error_log.append(entry)

    if not error_log:
        return

    df_err   = pd.DataFrame(error_log)
    sum_cols = ["Model", "Answer-type", "Error_Type"]
    if "target" in df_err.columns:
        sum_cols.append("target")
    summary  = df_err.groupby(sum_cols).size().reset_index(name="Count")

    summary.to_csv(out_dir / "error_summary.csv",    index=False, encoding="utf-8-sig")
    df_err.to_csv( out_dir / "error_detail_log.csv", index=False, encoding="utf-8-sig")
    print("    -> error_summary.csv  /  error_detail_log.csv")


def _ordered_model_columns(all_models: set) -> list:
    """Return a sorted column list: baselines first, then known models, then remaining."""
    base_cols = sorted(c for c in all_models if c.startswith("base_"))
    ordered   = [m for m in CUSTOM_MODEL_ORDER if m in all_models]
    remaining = sorted(all_models - set(base_cols) - set(ordered))
    return [c for c in (base_cols + ordered + remaining) if c in all_models]


def save_hierarchical_results(
    results_df: pd.DataFrame,
    label: str,
    out_dir: Path,
) -> None:
    """Write hierarchical evaluation results to eval_hierarchical_{label}.xlsx."""
    if results_df.empty:
        print(f"  No results for '{label}'.")
        return

    fname = out_dir / f"eval_hierarchical_{label}.xlsx"
    _TARGET_METRICS = [
        "F1-Macro", "RNRMSE_by_Range", "NRMSE_by_Size",
        "Format_Validity_F1", "Format_Validity_MSE", "Format_Validity_NRMSE",
    ]
    _CSV_METRICS = ["F1-Macro", "RNRMSE_by_Range", "NRMSE_by_Size"]

    final_cols_ref = [None]

    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        valid_metrics = [m for m in sorted(results_df["Metric"].unique())
                         if m in _TARGET_METRICS]

        for metric in valid_metrics:
            print(f"    Metric: {metric}  [{label}]")
            mdf = results_df[
                (results_df["Metric"] == metric) &
                (results_df["Hierarchy_Level"] != "Overall (No Size)")
            ].copy()
            if mdf.empty:
                continue

            try:
                p_vals = pd.pivot_table(
                    mdf, values="Value", index="Hierarchy",
                    columns="Model", aggfunc="mean",
                )

                if metric.startswith("Format_Validity"):
                    mdf["FmtVal"] = mdf.apply(
                        lambda r: f"{int(round(r['Value'] * r['Total']))}/{int(r['Total'])}",
                        axis=1,
                    )
                else:
                    mdf["FmtVal"] = mdf.apply(
                        lambda r: f"{r['Value']:.4f} ({int(r['Valid_Count'])}/{int(r['Total'])})",
                        axis=1,
                    )
                p_fmt = pd.pivot_table(
                    mdf, values="FmtVal", index="Hierarchy",
                    columns="Model", aggfunc="first",
                )

                final_cols        = _ordered_model_columns(set(p_vals.columns))
                final_cols_ref[0] = final_cols
                p_vals = p_vals[final_cols]
                p_fmt  = p_fmt[final_cols]

                def _nat_key(s: str) -> str:
                    s = str(s).lower()
                    s = re.sub(r"size_(\d+)",     lambda x: f"size_{x.group(1).zfill(6)}", s)
                    s = re.sub(r"instance_(\d+)", lambda x: f"instance_{x.group(1).zfill(6)}", s)
                    return s

                def _fmt_index(idx):
                    out = []
                    for label_str in idx:
                        parts  = label_str.split(" | ")
                        indent = "  " * (len(parts) - 1)
                        clean  = parts[-1]
                        suffix = " (all)" if "size_" in clean and len(parts) == 2 else ""
                        out.append(f"{indent}- {clean}{suffix}")
                    return out

                sorted_idx   = sorted(p_vals.index, key=_nat_key)
                p_vals       = p_vals.loc[sorted_idx]
                p_fmt        = p_fmt.loc[sorted_idx]
                p_vals.index = _fmt_index(p_vals.index)
                p_fmt.index  = _fmt_index(p_fmt.index)

                p_vals.round(4).to_excel(writer, sheet_name=f"{metric}_Values")
                p_fmt.to_excel(         writer, sheet_name=f"{metric}_Formatted")

                if metric in _CSV_METRICS:
                    safe = metric.replace(" ", "_").replace("/", "-")
                    csv_path = out_dir / f"metric_{safe}_{label}.csv"
                    p_vals.to_csv(csv_path, encoding="utf-8-sig")
                    print(f"      -> {csv_path.name}")

            except Exception as exc:
                print(f"    Error on metric '{metric}': {exc}")

        ov_df = results_df[results_df["Hierarchy_Level"] == "Overall (No Size)"].copy()
        if not ov_df.empty and final_cols_ref[0]:
            pivot_ov = pd.pivot_table(
                ov_df, values="Value",
                index=["Dataset", "Question-Type", "Metric"],
                columns="Model", aggfunc="mean",
            )
            cols = [c for c in final_cols_ref[0] if c in pivot_ov.columns]
            pivot_ov[cols].round(4).to_excel(writer, sheet_name="Overall_Performance")

    if writer.book.worksheets:
        print(f"  -> {fname.name}")
    elif fname.exists():
        fname.unlink()


def generate_category_summary(long_df: pd.DataFrame, out_dir: Path) -> None:
    """
    Compute F1 / RNRMSE / NRMSE per (model, dataset-category) and save
    category_summary.csv / category_summary.xlsx.
    """
    print("  Generating category summary ...")
    df = long_df.copy()
    df["Category"] = df["Dataset"].apply(get_dataset_category)
    df = df.dropna(subset=["Category"])
    if df.empty:
        actual = sorted(long_df["Dataset"].dropna().unique().tolist())
        print(f"    No matching categories. Actual Dataset values: {actual}")
        return

    records = []
    for (cat, model), grp in df.groupby(["Category", "Model"]):
        f1_val = rnrmse_val = nrmse_val = np.nan
        for atype, metric_name in [
            ("yes/no",     "F1-Macro"),
            ("percentage", "RNRMSE_by_Range"),
            ("number",     "NRMSE_by_Size"),
        ]:
            sub = grp[grp["Answer-type"] == atype]
            if sub.empty:
                continue
            res = calculate_metrics(sub, atype)
            if not res.empty:
                r = res[res["Metric"] == metric_name]
                if not r.empty:
                    v = r.iloc[0]["Value"]
                    if metric_name == "F1-Macro":          f1_val     = v
                    elif metric_name == "RNRMSE_by_Range": rnrmse_val = v
                    elif metric_name == "NRMSE_by_Size":   nrmse_val  = v

        records.append({"Model": model, "Category": cat,
                         "F1": f1_val, "RNRMSE": rnrmse_val, "NRMSE": nrmse_val})

    if not records:
        print("    No valid metrics."); return

    df_sum = pd.DataFrame(records)
    pivot  = df_sum.pivot(index="Model", columns="Category",
                           values=["F1", "RNRMSE", "NRMSE"])
    desired = [
        (met, cat)
        for cat in CAT_ORDER
        for met in ["F1", "RNRMSE", "NRMSE"]
        if (met, cat) in pivot.columns
    ]
    pivot    = pivot[desired]
    existing = [m for m in CUSTOM_MODEL_ORDER if m in pivot.index]
    others   = [m for m in pivot.index if m not in existing]
    pivot    = pivot.reindex(existing + others)

    print(pivot.to_string(float_format="%.4f", na_rep="-"))
    pivot.to_csv(  out_dir / "category_summary.csv")
    pivot.to_excel(out_dir / "category_summary.xlsx")
    print("    -> category_summary.csv  /  category_summary.xlsx")


# =============================================================================
# PART 2 – AGGREGATION  (M1 / Task Average only)
# =============================================================================

def _proc_metrics_agg(group: pd.DataFrame, answer_type_raw: str) -> dict:
    """
    Compute F1 / RNRMSE_by_Range / NRMSE for one answer-type slice.
    Used by generate_per_dataset_aggregation (M1 per-dataset, then averaged).
    Always uses strict error mode.
    """
    atype = str(answer_type_raw).lower().strip()
    y_true_cat, y_pred_cat = [], []
    y_true_num, y_pred_num, sizes, pct_ranges = [], [], [], []

    if atype == "percentage" and "Data_Size" in group.columns:
        _gold_by_size: dict = {}
        for _, _r in group.iterrows():
            _ds = int(_r.get("Data_Size", 0))
            _gv = to_numeric(_r["Gold_Label"])
            if pd.notna(_gv):
                _gold_by_size.setdefault(_ds, []).append(_gv)
        _size_ranges = {
            ds: (max(vs) - min(vs)) if len(vs) > 1 else PCT_GOLD_RANGE_FALLBACK
            for ds, vs in _gold_by_size.items()
        }
    else:
        _size_ranges = {}

    for _, row in group.iterrows():
        gold, out = row["Gold_Label"], row["Model_Output"]
        parsed    = parse_output(str(out), atype)
        data_size = int(row.get("Data_Size", 0))
        is_valid  = _validate_parsed(parsed, atype, data_size, gold)

        if is_valid:
            if atype in ("yes/no", "bool", "boolean"):
                y_true_cat.append(str(gold).lower().strip())
                y_pred_cat.append(parsed)
            else:
                y_true_num.append(to_numeric(gold))
                y_pred_num.append(parsed)
                if atype == "number":
                    sizes.append(data_size)
                elif atype == "percentage":
                    pct_ranges.append(_size_ranges.get(data_size, PCT_GOLD_RANGE_FALLBACK))
        else:
            # Strict penalty
            if atype in ("yes/no", "bool", "boolean"):
                g = str(gold).lower().strip()
                y_true_cat.append(g)
                y_pred_cat.append("no" if g == "yes" else "yes")
            elif atype in ("percentage", "number"):
                gn = to_numeric(gold)
                if pd.notna(gn):
                    y_true_num.append(gn)
                    if atype == "percentage":
                        y_pred_num.append(1.0 if gn <= 0.5 else 0.0)
                        pct_ranges.append(_size_ranges.get(data_size, PCT_GOLD_RANGE_FALLBACK))
                    else:
                        y_pred_num.append(float(data_size) if gn <= data_size / 2 else 0.0)
                        sizes.append(data_size)

    res = {"F1": np.nan, "RNRMSE_by_Range": np.nan, "NRMSE": np.nan}
    if y_true_cat:
        res["F1"] = float(f1_score(y_true_cat, y_pred_cat, average="macro",
                                    labels=["yes", "no"], zero_division=0))
    if y_true_num:
        if atype == "percentage":
            _ranges = pct_ranges if pct_ranges else [PCT_GOLD_RANGE_FALLBACK] * len(y_true_num)
            rnrmse_sq = [((t - p) / r) ** 2 if r > 0 else 0.0
                         for t, p, r in zip(y_true_num, y_pred_num, _ranges)]
            res["RNRMSE_by_Range"] = float(np.sqrt(np.mean(rnrmse_sq)))
        elif atype == "number" and sizes:
            sq = [((t - p) / s) ** 2 if s > 0 else 0.0
                  for t, p, s in zip(y_true_num, y_pred_num, sizes)]
            res["NRMSE"] = float(np.sqrt(np.mean(sq)))
    return res


def generate_per_dataset_aggregation(long_df: pd.DataFrame, out_dir: Path) -> None:
    """
    M1 (Task Average) per-dataset breakdown.

    For each (model, dataset), metrics are computed per answer-type slice,
    then averaged — preventing large datasets from dominating the aggregate.

    Output: aggr_per_dataset.xlsx  (Normal sheet + Transposed sheet)
      Normal     : rows = model groups, columns = dataset × metric
      Transposed : rows = dataset × metric, columns = models
    """
    from openpyxl.styles import Font

    METRICS        = ["F1", "RNRMSE_by_Range", "NRMSE"]
    METRIC_DISPLAY = ["F1", "RNRMSE", "NRMSE"]

    actual_ds = set(long_df["Dataset"].unique())
    # Preserve category ordering; fall back to sorted list if no categories match
    DS_ORDER = [kw for cat in CAT_ORDER for kw in DATASET_CATEGORIES[cat] if kw in actual_ds]
    if not DS_ORDER:
        DS_ORDER = sorted(actual_ds)
        print("    [WARN] No recognised dataset names — using all datasets in sorted order.")

    # ── Compute M1 per (model, dataset) ──────────────────────────────────────
    stats: list[dict] = []
    for (model, dataset), grp in long_df.groupby(["Model", "Dataset"]):
        row: dict = {"Model": model, "Dataset": dataset,
                     "F1": np.nan, "RNRMSE_by_Range": np.nan, "NRMSE": np.nan}
        for atype, sub in grp.groupby("Answer-type"):
            m = _proc_metrics_agg(sub, str(atype))
            for k in METRICS:
                if not np.isnan(m.get(k, np.nan)):
                    row[k] = m[k]
        stats.append(row)

    if not stats:
        print("    No metrics for per-dataset aggregation."); return

    stats_df    = pd.DataFrame(stats)
    base_mask   = stats_df["Model"].str.startswith("base_")
    real_models = sorted(stats_df.loc[~base_mask, "Model"].unique(), key=get_model_sort_key)
    base_models = sorted(stats_df.loc[base_mask,  "Model"].unique())
    model_order = real_models + base_models

    n_ds        = len(DS_ORDER)
    n_metrics   = len(METRICS)
    n_data_cols = n_ds * n_metrics

    def _val(model, dataset, metric):
        r = stats_df[(stats_df["Model"] == model) & (stats_df["Dataset"] == dataset)]
        if r.empty: return "-"
        v = r.iloc[0][metric]
        return f"{v:.4f}" if pd.notna(v) else "-"

    # Category spans for row-1 header
    cat_spans: list[tuple[str, int]] = []
    for cat in CAT_ORDER:
        n = sum(1 for kw in DATASET_CATEGORIES.get(cat, []) if kw in actual_ds)
        if n:
            cat_spans.append((cat, n * n_metrics))
    if not cat_spans:
        cat_spans = [("All Datasets", n_ds * n_metrics)]

    hdr1 = ["LLMs"] + [name for name, span in cat_spans for _ in range(span)]
    hdr2 = [""]     + [ds   for ds in DS_ORDER          for _ in range(n_metrics)]
    hdr3 = [""]     + METRIC_DISPLAY * n_ds
    data_rows: list[list] = []

    def _append_model_rows(model_list, with_type_header):
        current_type = None
        for model in model_list:
            mtype = get_model_type(model)
            if with_type_header and mtype != current_type:
                data_rows.append([mtype] + [""] * n_data_cols)
                current_type = mtype
            vals = [_val(model, ds, metric) for ds in DS_ORDER for metric in METRICS]
            data_rows.append([f"    {model}"] + vals)

    _append_model_rows(real_models, with_type_header=True)
    data_rows.append(["── Baselines ──"] + [""] * n_data_cols)
    _append_model_rows(base_models, with_type_header=False)

    # Transposed sheet: rows = (dataset, metric), columns = models
    trans_hdr  = ["Dataset", "Metric"] + [f"    {m}" for m in model_order]
    trans_rows = [trans_hdr]
    for ds in DS_ORDER:
        for metric, m_disp in zip(METRICS, METRIC_DISPLAY):
            trans_rows.append([ds, m_disp] + [_val(m, ds, metric) for m in model_order])

    # ── Write Excel ───────────────────────────────────────────────────────────
    fname = out_dir / "aggr_per_dataset.xlsx"
    with pd.ExcelWriter(fname, engine="openpyxl") as writer:
        wb = writer.book

        # Normal sheet
        ws_n = wb.create_sheet("Normal")
        for r_idx, row in enumerate([hdr1, hdr2, hdr3] + data_rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws_n.cell(row=r_idx, column=c_idx, value=val)

        # Merge category header cells (row 1)
        col_cursor = 2
        for cat_name, span in cat_spans:
            if span > 1:
                ws_n.merge_cells(start_row=1, start_column=col_cursor,
                                 end_row=1, end_column=col_cursor + span - 1)
            ws_n.cell(row=1, column=col_cursor).alignment = Alignment(
                horizontal="center", vertical="center")
            ws_n.cell(row=1, column=col_cursor).font = Font(bold=True)
            col_cursor += span

        # Merge dataset header cells (row 2)
        col_cursor = 2
        for _ in DS_ORDER:
            if n_metrics > 1:
                ws_n.merge_cells(start_row=2, start_column=col_cursor,
                                 end_row=2, end_column=col_cursor + n_metrics - 1)
            ws_n.cell(row=2, column=col_cursor).alignment = Alignment(
                horizontal="center", vertical="center")
            col_cursor += n_metrics

        # Centre metric labels (row 3)
        for c_idx in range(2, 2 + n_data_cols):
            ws_n.cell(row=3, column=c_idx).alignment = Alignment(horizontal="center")

        # Transposed sheet
        ws_t = wb.create_sheet("Transposed")
        for r_idx, row in enumerate(trans_rows, 1):
            for c_idx, val in enumerate(row, 1):
                ws_t.cell(row=r_idx, column=c_idx, value=val)
        for c_idx in range(1, len(trans_hdr) + 1):
            ws_t.cell(row=1, column=c_idx).font = Font(bold=True)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    print("  -> aggr_per_dataset.xlsx  (Normal + Transposed sheets)")


# =============================================================================
# TOP-LEVEL PIPELINE FUNCTIONS
# =============================================================================

def run_evaluation(file_path: str, results_dir: str = None) -> None:
    """
    Part 1: hierarchical per-dataset evaluation.

    Reads any CSV file containing the required columns.
    Output directory defaults to <input_stem>_eval/ next to the input file.
    """
    try:
        df = pd.read_csv(file_path)
    except FileNotFoundError:
        print(f"Error: file not found – {file_path}"); return

    out_dir = (Path(results_dir) if results_dir
               else Path(file_path).parent / f"{Path(file_path).stem}_eval")
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*64}")
    print(f"  EVALUATION  |  {Path(file_path).name}")
    print(f"  Output     : {out_dir}")
    print(f"{'='*64}")

    if "target" in df.columns:
        df["target"] = df["target"].astype(str).replace("nan", np.nan)

    try:
        long_df = transform_wide_to_long(df)
    except ValueError as e:
        print(f"  Error: {e}"); return

    if long_df.empty:
        print("  No data after transformation."); return

    long_df = add_baselines(long_df)
    save_error_report(long_df, out_dir)

    if "Question Category" in long_df.columns:
        df_general  = long_df[long_df["Question Category"].str.lower() == "general"].copy()
        df_datasets = long_df[long_df["Question Category"].str.lower() != "general"].copy()
    else:
        df_general  = pd.DataFrame()
        df_datasets = long_df.copy()

    if not df_general.empty:
        print("\n  [General questions]")
        res = perform_hierarchical_evaluation(df_general)
        save_hierarchical_results(res, "general", out_dir)

    if not df_datasets.empty:
        print("\n  [Dataset questions]")
        res = perform_hierarchical_evaluation(df_datasets)
        save_hierarchical_results(res, "datasets", out_dir)

    generate_category_summary(long_df, out_dir)
    print(f"\n  Evaluation complete -> {out_dir}\n")


def run_aggregation(file_path: str, results_dir: str = None) -> None:
    """
    Part 2: M1 (Task Average) per-dataset aggregation.
    Outputs only aggr_per_dataset.xlsx.
    """
    in_path = Path(file_path)
    out_dir = Path(results_dir) if results_dir else in_path.parent / f"{in_path.stem}_aggr"
    out_dir.mkdir(parents=True, exist_ok=True)

    if not in_path.exists():
        print(f"Error: file not found – {in_path}"); return

    print(f"\n{'='*64}")
    print(f"  AGGREGATION  |  {in_path.name}")
    print(f"  Output      : {out_dir}")
    print(f"{'='*64}")

    df = pd.read_csv(in_path)
    try:
        long_df = transform_wide_to_long(df)
    except ValueError as e:
        print(f"  Error: {e}"); return

    if long_df.empty:
        print("  No data after transformation."); return

    long_df = add_baselines(long_df)
    print("  Computing M1 (Task Average) per-dataset aggregation ...")
    generate_per_dataset_aggregation(long_df, out_dir)
    print(f"\n  Aggregation complete -> {out_dir}\n")


def run_full_pipeline(
    file_path: str,
    eval_dir: str = None,
    aggr_dir: str = None,
) -> None:
    """Run Part 1 (evaluation) then Part 2 (aggregation) for one input CSV."""
    run_evaluation(file_path, eval_dir)
    run_aggregation(file_path, aggr_dir)


# =============================================================================
# ENTRY POINT
# =============================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description=(
            "Evaluation and aggregation pipeline for LLM text-analysis benchmarks.\n\n"
            "Required CSV columns : Dataset, Question-Type, Answer-type, Question\n"
            "Gold columns         : gold_{size}_{sample}  (any subset of sizes/samples)\n"
            "Model columns        : res_{size}_{sample}_{model_name}\n"
            "Baseline columns     : base_{name}  (e.g. base_yes, base_0%, base_100)"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--input", required=True, metavar="CSV",
        help="Path to the wide-format input CSV file (any filename accepted).",
    )
    parser.add_argument(
        "--eval_dir", default=None, metavar="DIR",
        help=(
            "Output directory for evaluation results. "
            "Default: <input_stem>_eval/ next to the input file."
        ),
    )
    parser.add_argument(
        "--aggr_dir", default=None, metavar="DIR",
        help=(
            "Output directory for aggregation results. "
            "Default: <input_stem>_aggr/ next to the input file."
        ),
    )
    parser.add_argument(
        "--eval_only", action="store_true",
        help="Run only Part 1 (hierarchical evaluation); skip aggregation.",
    )
    parser.add_argument(
        "--aggr_only", action="store_true",
        help="Run only Part 2 (per-dataset aggregation); skip evaluation.",
    )
    args = parser.parse_args()

    if args.eval_only and args.aggr_only:
        parser.error("--eval_only and --aggr_only are mutually exclusive.")

    if args.eval_only:
        run_evaluation(args.input, args.eval_dir)
    elif args.aggr_only:
        run_aggregation(args.input, args.aggr_dir)
    else:
        run_full_pipeline(args.input, args.eval_dir, args.aggr_dir)
